from fastapi import FastAPI, Query, Depends, APIRouter
from pydantic import BaseModel
from typing import List, Optional
from datetime import date, datetime
from sqlalchemy.orm import Session
from app.models import Ticket, SeatLog, Seat, Counter, Tenxa  # assuming these are your SQLAlchemy models
from sqlalchemy import func, and_, or_
from app import crud, schemas, database
from collections import defaultdict
from datetime import datetime, timedelta, time
import pytz
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font

#app = FastAPI()
router = APIRouter()
def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_stats_db():
    db = database.SessionStats()
    try:
        yield db
    finally:
        db.close()


# ==== SCHEMAS ====

class TicketsPerCounter(BaseModel):
    counter_id: int
    total_tickets: int

class AttendedTickets(BaseModel):
    counter_id: int
    attended_tickets: int

class AverageHandlingTime(BaseModel):
    counter_id: int
    avg_handling_time_seconds: float

class AbsenceTime(BaseModel):
    counter_id: int
    total_absence_seconds: float

class WorkingTimeCheck(BaseModel):
    counter_id: int
    is_late: bool
    first_checkin: Optional[datetime]
    
class AfkDuration(BaseModel):
    counter_id: int
    total_absent_minutes: float

class AverageWaitingTime(BaseModel):
    counter_id: int
    avg_waiting_time_seconds: float


# ==== UTILS ====

def get_date_range(start: Optional[date], end: Optional[date]):
    today = date.today()
    if not start:
        start = today
    if not end:
        end = today
    return start, end


# ==== ENDPOINTS ====

@router.get("/tickets-per-counter", response_model=List[TicketsPerCounter])
def tickets_per_counter(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    print("start_date:", start_date, "end_date:", end_date)
    start, end = get_date_range(start_date, end_date)

    result = (
        db.query(Ticket.counter_id, func.count().label("total_tickets"))
        .filter(func.date(Ticket.created_at) >= start, func.date(Ticket.created_at) <= end)
        .filter(Ticket.tenxa_id == tenxa_id)
        .group_by(Ticket.counter_id)
        .all()
    )

    items = [
        TicketsPerCounter(counter_id=row[0], total_tickets=row[1])
        for row in result
    ]

    # ✅ Tính tổng tất cả vé
    total = sum(row[1] for row in result)
    #items.append(TicketsPerCounter(counter_id="Tổng", total_tickets=total))

    return items


@router.get("/attended-tickets", response_model=List[AttendedTickets])
def attended_tickets(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    start, end = get_date_range(start_date, end_date)

    result = (
        db.query(Ticket.counter_id, func.count().label("attended_tickets"))
        .filter(
            Ticket.called_at.isnot(None),
            Ticket.finished_at.isnot(None),
            func.date(Ticket.created_at) >= start,
            func.date(Ticket.created_at) <= end,
        )
        .filter(Ticket.tenxa_id == tenxa_id)
        .group_by(Ticket.counter_id)
        .all()
    )

    items = [
        AttendedTickets(counter_id=row[0], attended_tickets=row[1])
        for row in result
    ]

    # ✅ Thêm dòng tổng
    total = sum(row[1] for row in result)
    #items.append(AttendedTickets(counter_id="Tổng", attended_tickets=total))

    return items



@router.get("/average-handling-time", response_model=List[AverageHandlingTime])
def average_handling_time(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    start, end = get_date_range(start_date, end_date)
    result = (
        db.query(
            Ticket.counter_id,
            func.avg(func.extract("epoch", Ticket.finished_at - Ticket.called_at)).label("avg_handling_time_minutes")
        )
        .filter(Ticket.tenxa_id == tenxa_id)
        .filter(
            Ticket.called_at.isnot(None),
            Ticket.finished_at.isnot(None),
            func.date(Ticket.created_at) >= start,
            func.date(Ticket.created_at) <= end,
        )
        .group_by(Ticket.counter_id)
        .all()
    )

    return [
        AverageHandlingTime(counter_id=row[0], avg_handling_time_seconds=row[1])
        for row in result
    ]


@router.get("/working-time-check", response_model=List[WorkingTimeCheck])
def working_time_check(
    date_check: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    date_check = date_check or date.today()
    result = []

    sub = (
        db.query(
            Seat.counter_id,
            func.min(SeatLog.timestamp).label("first_checkin")
        )
        .join(Seat, Seat.id == SeatLog.seat_id)
        .filter(SeatLog.tenxa_id == tenxa_id)
        .filter(
            SeatLog.new_status == True,  # Có mặt
            func.date(SeatLog.timestamp) == date_check
        )
        .group_by(Seat.counter_id)
        .all()
    )

    for counter_id, first_checkin in sub:
        is_late = first_checkin.time() > datetime.strptime("07:30:00", "%H:%M:%S").time()
        result.append(
            WorkingTimeCheck(
                counter_id=counter_id,
                is_late=is_late,
                first_checkin=first_checkin
            )
        )

    return result

@router.get("/afk-duration", response_model=List[AfkDuration])
def afk_duration(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)

    start_date, end_date = get_date_range(start_date, end_date)
    total_afk_per_counter = defaultdict(float)

    seat_logs = (
        db.query(SeatLog)
        .join(Seat, SeatLog.seat_id == Seat.id)
        .filter(SeatLog.tenxa_id == tenxa_id)
        .filter(
            func.date(SeatLog.timestamp) >= start_date,
            func.date(SeatLog.timestamp) <= end_date,
            SeatLog.new_status.in_([True, False])  # 0: vắng mặt, 1: có mặt
        )
        .order_by(SeatLog.seat_id, SeatLog.timestamp)
        .all()
    )

    from itertools import groupby
    grouped_by_seat = groupby(seat_logs, key=lambda log: log.seat_id)

    for seat_id, logs in grouped_by_seat:
        logs = list(logs)
        counter_id = logs[0].seat.counter_id
        prev_status = None
        prev_time = None

        for log in logs:
            current_time = log.timestamp
            current_status = log.new_status

            if prev_status is False and current_status is True:
                afk_start = prev_time
                afk_end = current_time

                tz = afk_start.tzinfo or pytz.timezone("Asia/Ho_Chi_Minh")

                working_start = tz.localize(datetime.combine(afk_start.date(), time(7, 30))) if afk_start.tzinfo is None else datetime.combine(afk_start.date(), time(7, 30)).replace(tzinfo=tz)
                working_end = tz.localize(datetime.combine(afk_end.date(), time(17, 30))) if afk_end.tzinfo is None else datetime.combine(afk_end.date(), time(17, 30)).replace(tzinfo=tz)

                effective_start = max(afk_start, working_start)
                effective_end = min(afk_end, working_end)

                if effective_start < effective_end:
                    duration = (effective_end - effective_start).total_seconds()
                    total_afk_per_counter[counter_id] += duration / 60  # convert to minutes

            prev_status = current_status
            prev_time = current_time

    return [
        AfkDuration(counter_id=k, total_absent_minutes=v)
        for k, v in total_afk_per_counter.items()
    ]

@router.get("/average-waiting-time", response_model=List[AverageWaitingTime])
def average_waiting_time(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    start, end = get_date_range(start_date, end_date)

    result = (
        db.query(
            Ticket.counter_id,
            func.avg(func.extract("epoch", Ticket.called_at - Ticket.created_at)).label("avg_waiting_time_minutes")
        )
        .filter(
            Ticket.created_at.isnot(None),
            Ticket.called_at.isnot(None),
            func.date(Ticket.created_at) >= start,
            func.date(Ticket.created_at) <= end,
        )
        .filter(Ticket.tenxa_id == tenxa_id)
        .group_by(Ticket.counter_id)
        .all()
    )

    return [
        AverageWaitingTime(counter_id=row[0], avg_waiting_time_seconds=row[1])
        for row in result
    ]

@router.get("/export/ticket-report")
def export_ticket_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    format: str = Query("excel", regex="^(excel|pdf)$"),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    start, end = get_date_range(start_date, end_date)

    # Query vé đã tiếp đón
    tickets = (
        db.query(Ticket)
        .filter(
            Ticket.tenxa_id == tenxa_id,
            Ticket.called_at.isnot(None),
            Ticket.finished_at.isnot(None),
            func.date(Ticket.created_at) >= start,
            func.date(Ticket.created_at) <= end,
        )
        .order_by(Ticket.created_at)
        .all()
    )

    # Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ticket Report"

    headers = [
        "Quầy",
        "Số vé",
        "Thời điểm in vé",
        "Thời điểm tiếp đón",
        "Tổng thời gian tiếp đón (phút)",
        "Cảnh báo (>15 phút)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for ticket in tickets:
        duration_minutes = (ticket.finished_at - ticket.called_at).total_seconds() / 60
        warning = "⚠️" if duration_minutes > 15 else ""
        ws.append([
            ticket.counter_id,
            ticket.number,
            ticket.created_at.strftime("%Y-%m-%d %H:%M"),
            ticket.called_at.strftime("%Y-%m-%d %H:%M"),
            round(duration_minutes, 2),
            warning,
        ])

    # Trả file về client
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ticket_report_{tenxa}_{start}_{end}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

class RatingPerCounter(BaseModel):
    counter_id: int
    satisfied: int
    neutral: int
    need_improvement: int

@router.get("/rating-per-counter", response_model=List[RatingPerCounter])
def rating_per_counter(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db),
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    start, end = get_date_range(start_date, end_date)

    result = (
        db.query(
            Ticket.counter_id,
            Ticket.rating,
            func.count().label("count")
        )
        .filter(
            Ticket.tenxa_id == tenxa_id,
            Ticket.rating.isnot(None),
            Ticket.status == "done",
            func.date(Ticket.created_at).between(start, end)
        )
        .group_by(Ticket.counter_id, Ticket.rating)
        .all()
    )

    # Gom theo counter
    data = defaultdict(lambda: {"satisfied": 0, "neutral": 0, "needs_improvement": 0})
    for counter_id, rating, count in result:
        data[counter_id][rating] = count

    return [
        RatingPerCounter(
            counter_id=cid,
            satisfied=vals["satisfied"],
            neutral=vals["neutral"],
            need_improvement=vals["needs_improvement"]
        )
        for cid, vals in data.items()
    ]

class FeedbackItem(BaseModel):
    ticket_number: int
    counter_id: int
    rating: str
    feedback: Optional[str] = None
    created_at: datetime

@router.get("/feedbacks", response_model=List[FeedbackItem])
def list_feedbacks(
    rating: Optional[str] = Query(None, regex="^(satisfied|neutral|needs_improvement)$"),
    counter_id: Optional[int] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    tenxa: str = Query(...),
    db: Session = Depends(get_db)
):
    tenxa_id = crud.get_tenxa_id_from_slug(db, tenxa)
    start, end = get_date_range(start_date, end_date)

    q = db.query(Ticket).filter(
        Ticket.tenxa_id == tenxa_id,
        #Ticket.feedback.isnot(None),
        Ticket.status == "done",
        func.date(Ticket.created_at).between(start, end)
    )
    if rating:
        q = q.filter(Ticket.rating == rating)
    if counter_id:
        q = q.filter(Ticket.counter_id == counter_id)

    tickets = q.order_by(Ticket.created_at.desc()).all()

    return [
        FeedbackItem(
            ticket_number=t.number,
            counter_id=t.counter_id,
            rating=t.rating,
            feedback=t.feedback,
            created_at=t.created_at
        )
        for t in tickets
    ]

class TenxaStats(BaseModel):
    tenxa_id: int
    tenxa_name: str
    total_tickets: int
    attended_tickets: int
    avg_waiting_time_seconds: Optional[float]
    avg_handling_time_seconds: Optional[float]
    satisfied: int
    neutral: int
    needs_improvement: int

@router.get("/all-unit", response_model=List[TenxaStats])
def stats_by_tenxa(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_stats_db),
):
    start, end = get_date_range(start_date, end_date)

    # --- Tổng số vé được in ---
    total_q = (
        db.query(
            Ticket.tenxa_id,
            func.count().label("total_tickets")
        )
        .filter(func.date(Ticket.created_at).between(start, end))
        .group_by(Ticket.tenxa_id)
        .all()
    )
    total_map = {row.tenxa_id: row.total_tickets for row in total_q}

    # --- Tổng số vé đã tiếp đón ---
    attended_q = (
        db.query(
            Ticket.tenxa_id,
            func.count().label("attended_tickets")
        )
        .filter(
            Ticket.called_at.isnot(None),
            Ticket.finished_at.isnot(None),
            func.date(Ticket.created_at).between(start, end)
        )
        .group_by(Ticket.tenxa_id)
        .all()
    )
    attended_map = {row.tenxa_id: row.attended_tickets for row in attended_q}

    # --- Thời gian chờ trung bình ---
    waiting_q = (
        db.query(
            Ticket.tenxa_id,
            func.avg(func.extract("epoch", Ticket.called_at - Ticket.created_at)).label("avg_waiting_time")
        )
        .filter(
            Ticket.created_at.isnot(None),
            Ticket.called_at.isnot(None),
            func.date(Ticket.created_at).between(start, end)
        )
        .group_by(Ticket.tenxa_id)
        .all()
    )
    waiting_map = {row.tenxa_id: row.avg_waiting_time for row in waiting_q}

    # --- Thời gian tiếp đón trung bình ---
    handling_q = (
        db.query(
            Ticket.tenxa_id,
            func.avg(func.extract("epoch", Ticket.finished_at - Ticket.called_at)).label("avg_handling_time")
        )
        .filter(
            Ticket.called_at.isnot(None),
            Ticket.finished_at.isnot(None),
            func.date(Ticket.created_at).between(start, end)
        )
        .group_by(Ticket.tenxa_id)
        .all()
    )
    handling_map = {row.tenxa_id: row.avg_handling_time for row in handling_q}

    # --- Rating ---
    rating_q = (
        db.query(
            Ticket.tenxa_id,
            Ticket.rating,
            func.count().label("count")
        )
        .filter(
            Ticket.rating.isnot(None),
            Ticket.status == "done",   # 👈 chỉ lấy vé done
            func.date(Ticket.created_at).between(start, end)
        )
        .group_by(Ticket.tenxa_id, Ticket.rating)
        .all()
    )


    rating_map = defaultdict(lambda: {"satisfied": 0, "neutral": 0, "needs_improvement": 0})
    for tenxa_id, rating, count in rating_q:
        if rating in rating_map[tenxa_id]:
            rating_map[tenxa_id][rating] += count

    # --- Lấy danh sách xã ---
    tenxa_list = db.query(Tenxa.id, Tenxa.name).all()

    results = []
    for tx_id, tx_name in tenxa_list:
        results.append(
            TenxaStats(
                tenxa_id=tx_id,
                tenxa_name=tx_name,
                total_tickets=total_map.get(tx_id, 0),
                attended_tickets=attended_map.get(tx_id, 0),
                avg_waiting_time_seconds=waiting_map.get(tx_id),
                avg_handling_time_seconds=handling_map.get(tx_id),
                satisfied=rating_map[tx_id]["satisfied"],
                neutral=rating_map[tx_id]["neutral"],
                needs_improvement=rating_map[tx_id]["needs_improvement"],  # ⚠️ sửa đúng key
            )
        )

    return results


from fastapi import Query, Depends, APIRouter
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from statistics import mean

from app import database
from app.models import Ticket, Tenxa
from .your_stats_file import stats_by_tenxa  # 👈 thay đúng đường dẫn file chứa stats_by_tenxa


router = APIRouter()

def get_stats_db():
    db = database.StatsSessionLocal()  # 👈 nếu bạn dùng DB riêng cho thống kê
    try:
        yield db
    finally:
        db.close()


@router.get("/all-unit/excel")
def export_stats_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_stats_db),
):
    start, end = stats_by_tenxa.__globals__["get_date_range"](start_date, end_date)  # reuse util
    stats = stats_by_tenxa(start_date, end_date, db)

    # --- Sort theo mã xã ---
    stats_sorted = sorted(stats, key=lambda r: r.tenxa_id)

    # --- Tạo workbook ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Thống kê xã"

    # --- Style cơ bản ---
    bold_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="BDD7EE")  # xanh nhạt
    total_fill = PatternFill("solid", fgColor="FFD966")   # vàng nhạt
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Tiêu đề ---
    title = f"BÁO CÁO THỐNG KÊ THEO XÃ ({start} → {end})"
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(bold=True, size=14, color="1F4E78")
    cell.alignment = center_align

    # --- Để trống 1 dòng, header bắt đầu từ dòng 3 ---
    headers = [
        "Mã xã",
        "Tên xã",
        "Tổng vé",
        "Vé đã tiếp đón",
        "TG chờ TB (phút)",         # 👈 đổi thành phút
        "TG tiếp đón TB (phút)",    # 👈 đổi thành phút
        "Hài lòng",
        "Bình thường",
        "Cần cải thiện",
    ]
    ws.append([])  # dòng 2 trống
    ws.append(headers)

    # style header (dòng 3)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # --- Dữ liệu ---
    waiting_times = []
    handling_times = []
    for row in stats_sorted:
        waiting_min = (row.avg_waiting_time_seconds or 0) / 60
        handling_min = (row.avg_handling_time_seconds or 0) / 60

        ws.append([
            row.tenxa_id or 0,
            row.tenxa_name or "",
            row.total_tickets or 0,
            row.attended_tickets or 0,
            round(waiting_min, 2),
            round(handling_min, 2),
            row.satisfied or 0,
            row.neutral or 0,
            row.needs_improvement or 0,
        ])

        if row.avg_waiting_time_seconds:
            waiting_times.append(waiting_min)
        if row.avg_handling_time_seconds:
            handling_times.append(handling_min)

    # style dữ liệu
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # --- Thêm dòng tổng kết ---
    ws.append([])  # dòng trống
    total_row_idx = ws.max_row + 1

    ws.cell(row=total_row_idx, column=1, value="TỔNG KẾT")
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=2)

    ws.cell(row=total_row_idx, column=3, value=sum(r.total_tickets or 0 for r in stats_sorted))
    ws.cell(row=total_row_idx, column=4, value=sum(r.attended_tickets or 0 for r in stats_sorted))
    ws.cell(row=total_row_idx, column=5, value=round(mean(waiting_times), 2) if waiting_times else 0)
    ws.cell(row=total_row_idx, column=6, value=round(mean(handling_times), 2) if handling_times else 0)
    ws.cell(row=total_row_idx, column=7, value=sum(r.satisfied or 0 for r in stats_sorted))
    ws.cell(row=total_row_idx, column=8, value=sum(r.neutral or 0 for r in stats_sorted))
    ws.cell(row=total_row_idx, column=9, value=sum(r.needs_improvement or 0 for r in stats_sorted))

    # style dòng tổng kết
    for col in range(1, 10):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Xuất file ---
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"thong_ke_xa_{start}_{end}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

